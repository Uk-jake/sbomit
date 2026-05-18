#!/usr/bin/env python3
"""
aggregate_results.py — Roll up experiment results into an Excel report.

After run_batch.py (or experiment.py) has run, each project has an
experiments/<project>/result.json. This script reads all of them and produces
a single Excel workbook with two sheets:

  Sheet "Projects"  — one row per project. The coverage view: which projects
                      produced an SBOM, how many packages, build system, etc.
                      This is the answer to "how many projects can SBOMit
                      generate an SBOM for".

  Sheet "Steps"     — one row per step across all projects. The failure-analysis
                      view: which steps failed, their exit codes, and an empty
                      'failure_category' / 'failure_summary' column for you to
                      fill in by hand after reading the logs.

Deliberately NOT included (per the agreed plan):
  - sbom_from_build_only: needs the compare_steps analysis, which is on its own
    branch and still has the eBPF broken-path issue (76.8% of paths corrupted
    in heavy builds). Left for a second experiment round.

Usage:
  python3 aggregate_results.py                 # all projects in experiments/
  python3 aggregate_results.py --list projects.txt
  python3 aggregate_results.py --output report.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


SCRIPT_DIR      = Path(__file__).resolve().parent
EXPERIMENTS_DIR = SCRIPT_DIR / "experiments"
DEFAULT_OUTPUT  = SCRIPT_DIR / "sbomit_coverage_report.xlsx"

# Build-step heuristic — same hints as compare_steps.py, kept consistent.
BUILD_STEP_HINTS = ("build", "compile", "default", "install")

# Styling constants.
FONT_NAME    = "Arial"
HEADER_FILL  = PatternFill("solid", start_color="1F4E78")   # dark blue
HEADER_FONT  = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
CELL_FONT    = Font(name=FONT_NAME, size=10)
OK_FILL      = PatternFill("solid", start_color="C6EFCE")   # green
FAIL_FILL    = PatternFill("solid", start_color="FFC7CE")   # red
SKIP_FILL    = PatternFill("solid", start_color="F2F2F2")   # grey
INPUT_FILL   = PatternFill("solid", start_color="FFFF00")   # yellow — fill me in


def is_build_step(step_name: str) -> bool:
    """True if a step name looks like a build step (heuristic)."""
    low = step_name.lower()
    return any(hint in low for hint in BUILD_STEP_HINTS)


def find_result_files(project_filter: list[str] | None) -> list[Path]:
    """Locate experiments/<project>/result.json files.

    If project_filter is given, only those projects are included (and a missing
    one is reported). Otherwise every experiments/*/result.json is used.
    """
    if not EXPERIMENTS_DIR.is_dir():
        print(f"ERROR: no experiments/ directory at {EXPERIMENTS_DIR}",
              file=sys.stderr)
        sys.exit(2)

    if project_filter:
        files = []
        for name in project_filter:
            rj = EXPERIMENTS_DIR / name / "result.json"
            if rj.exists():
                files.append(rj)
            else:
                print(f"  WARNING: no result.json for '{name}' "
                      f"(experiment not run?)")
        return files

    return sorted(EXPERIMENTS_DIR.glob("*/result.json"))


def load_result(result_path: Path) -> dict | None:
    """Load one result.json, returning None on failure."""
    try:
        return json.loads(result_path.read_text())
    except Exception as e:
        print(f"  WARNING: could not read {result_path}: {e}")
        return None


# ──────────────────────────────────────────────────────────────────────────────
# Row extraction
# ──────────────────────────────────────────────────────────────────────────────
def project_row(result: dict) -> dict:
    """Build one Projects-sheet row from a result.json dict."""
    steps = result.get("steps", [])
    summary = result.get("summary", {})
    sbom_gen = result.get("sbom_generation", {})
    attestations = result.get("attestations", {})

    # Did any build-like step succeed?
    build_ok = any(
        is_build_step(s.get("name", "")) and s.get("status") == "ok"
        for s in steps
    )

    # SBOM generated if any format succeeded; record the max package count.
    sbom_formats_ok = [f for f, r in sbom_gen.items() if r.get("success")]
    sbom_generated = len(sbom_formats_ok) > 0
    sbom_packages = max(
        (r.get("packages", 0) for r in sbom_gen.values() if r.get("success")),
        default=0,
    )

    return {
        "project": result.get("project", "?"),
        "build_system": result.get("build_system") or "(none)",
        "total_steps": summary.get("total", 0),
        "ok_steps": summary.get("ok", 0),
        "failed_steps": summary.get("failed", 0),
        "skipped_steps": summary.get("skipped", 0),
        "build_step_succeeded": "Y" if build_ok else "N",
        "sbom_generated": "Y" if sbom_generated else "N",
        "sbom_packages": sbom_packages,
        "sbom_formats_ok": ", ".join(sorted(sbom_formats_ok)) or "-",
        "attestations_uploaded": attestations.get("uploaded", 0),
        "duration_s": result.get("duration_s", 0),
        "completed_at": result.get("completed_at", ""),
    }


def step_rows(result: dict) -> list[dict]:
    """Build Steps-sheet rows from a result.json dict (one per step)."""
    project = result.get("project", "?")
    rows = []
    for s in result.get("steps", []):
        status = s.get("status", "?")
        # exit_code only carries meaning for a failed step; for ok/skipped
        # steps it is noise in a failure-analysis sheet, so leave it blank.
        exit_code = s.get("exit_code", "") if status == "failed" else ""
        rows.append({
            "project": project,
            "step": s.get("name", "?"),
            "status": status,
            "is_build_step": "Y" if is_build_step(s.get("name", "")) else "N",
            "exit_code": exit_code,
            "log_path": s.get("log_path", "") or "",
            "failure_category": "",   # you fill: A=env / B=dependency / C=non-build
            "failure_summary": "",    # you fill: one-line cause from the log
        })
    return rows


# ──────────────────────────────────────────────────────────────────────────────
# Excel writing
# ──────────────────────────────────────────────────────────────────────────────
def _write_sheet(ws, headers: list[str], rows: list[dict],
                 input_columns: set[str] | None = None) -> None:
    """Write a header row + data rows to a worksheet with standard styling.

    input_columns: column keys that the user is expected to fill in by hand —
    highlighted yellow so they stand out as "needs attention".
    """
    input_columns = input_columns or set()

    # Header row.
    for col_idx, key in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows.
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(headers, start=1):
            value = row.get(key, "")
            # An empty / None value: write "" so the cell is visibly blank
            # rather than the literal text "None".
            if value is None or value == "":
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT

            # Highlight user-input columns.
            if key in input_columns:
                cell.fill = INPUT_FILL

            # Status-based row coloring cues.
            if key == "status":
                if value == "ok":
                    cell.fill = OK_FILL
                elif value == "failed":
                    cell.fill = FAIL_FILL
                elif value == "skipped":
                    cell.fill = SKIP_FILL
            if key == "sbom_generated":
                cell.fill = OK_FILL if value == "Y" else FAIL_FILL

    # Column widths — size to the longest cell, within reason.
    for col_idx, key in enumerate(headers, start=1):
        longest = len(str(key))
        for row in rows:
            longest = max(longest, len(str(row.get(key, ""))))
        ws.column_dimensions[get_column_letter(col_idx)].width = \
            min(max(longest + 2, 10), 50)

    # Freeze the header row.
    ws.freeze_panes = "A2"


def build_workbook(project_rows: list[dict],
                   all_step_rows: list[dict]) -> Workbook:
    """Assemble the two-sheet workbook."""
    wb = Workbook()

    # ── Sheet 1: Projects (coverage view) ─────────────────────────────────────
    ws_proj = wb.active
    ws_proj.title = "Projects"
    proj_headers = [
        "project", "build_system",
        "total_steps", "ok_steps", "failed_steps", "skipped_steps",
        "build_step_succeeded", "sbom_generated", "sbom_packages",
        "sbom_formats_ok", "attestations_uploaded",
        "duration_s", "completed_at",
    ]
    _write_sheet(ws_proj, proj_headers, project_rows)

    # A small summary block below the table — the headline coverage number.
    summary_row = len(project_rows) + 3
    total = len(project_rows)
    with_sbom = sum(1 for r in project_rows if r["sbom_generated"] == "Y")
    ws_proj.cell(row=summary_row, column=1, value="COVERAGE SUMMARY").font = \
        Font(name=FONT_NAME, bold=True, size=11)
    ws_proj.cell(row=summary_row + 1, column=1, value="Projects analyzed")
    ws_proj.cell(row=summary_row + 1, column=2, value=total)
    ws_proj.cell(row=summary_row + 2, column=1, value="SBOM generated")
    ws_proj.cell(row=summary_row + 2, column=2, value=with_sbom)
    ws_proj.cell(row=summary_row + 3, column=1, value="Coverage")
    # Formula, not a hardcoded value — recalculated by recalc.py.
    ws_proj.cell(row=summary_row + 3, column=2,
                 value=f"=B{summary_row + 2}/B{summary_row + 1}")
    ws_proj.cell(row=summary_row + 3, column=2).number_format = "0.0%"
    for r in range(summary_row + 1, summary_row + 4):
        ws_proj.cell(row=r, column=1).font = CELL_FONT
        ws_proj.cell(row=r, column=2).font = CELL_FONT

    # ── Sheet 2: Steps (failure-analysis view) ────────────────────────────────
    ws_steps = wb.create_sheet("Steps")
    step_headers = [
        "project", "step", "status", "is_build_step", "exit_code",
        "log_path", "failure_category", "failure_summary",
    ]
    # The last two columns are filled in by hand — highlight them yellow.
    _write_sheet(ws_steps, step_headers, all_step_rows,
                 input_columns={"failure_category", "failure_summary"})

    # A legend note for the manual columns.
    note_row = len(all_step_rows) + 3
    ws_steps.cell(row=note_row, column=1,
                  value="failure_category: A = environment, "
                        "B = build dependency, C = non-build step").font = \
        Font(name=FONT_NAME, italic=True, size=9)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Aggregate experiment result.json files into an Excel "
                    "report.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Example:\n  python3 aggregate_results.py --output report.xlsx",
    )
    parser.add_argument(
        "--list", default=None,
        help="Project list file; only these projects are aggregated. "
             "Default: every project found under experiments/.",
    )
    parser.add_argument(
        "--output", default=str(DEFAULT_OUTPUT),
        help=f"Output .xlsx path (default: {DEFAULT_OUTPUT.name}).",
    )
    args = parser.parse_args()

    project_filter = None
    if args.list:
        list_path = Path(args.list)
        if not list_path.exists():
            print(f"ERROR: list file not found: {list_path}", file=sys.stderr)
            sys.exit(2)
        project_filter = [
            ln.strip() for ln in list_path.read_text().splitlines()
            if ln.strip() and not ln.strip().startswith("#")
        ]

    result_files = find_result_files(project_filter)
    if not result_files:
        print("No result.json files found — run experiment.py / run_batch.py "
              "first.")
        sys.exit(1)

    print(f"Aggregating {len(result_files)} project result(s)...")

    project_rows: list[dict] = []
    all_step_rows: list[dict] = []
    for rf in result_files:
        result = load_result(rf)
        if result is None:
            continue
        project_rows.append(project_row(result))
        all_step_rows.extend(step_rows(result))
        print(f"  {result.get('project', '?')}: "
              f"{len(result.get('steps', []))} steps")

    if not project_rows:
        print("No readable results.")
        sys.exit(1)

    wb = build_workbook(project_rows, all_step_rows)
    out_path = Path(args.output)
    wb.save(out_path)

    print()
    print(f"Report written: {out_path}")
    print(f"  Projects sheet : {len(project_rows)} rows")
    print(f"  Steps sheet    : {len(all_step_rows)} rows")
    print(f"  (fill in failure_category / failure_summary by hand — "
          f"yellow columns)")


if __name__ == "__main__":
    main()